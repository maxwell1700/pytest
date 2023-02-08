from helper import validate_conversion
import django
import openpyxl
#if you dont want to import something specific you just import helper and use(helper.validate_convertsion())

#print(validate_conversion())


my_file = openpyxl.load_workbook('/Users/maxwelltrimnell/Desktop/pytest/inventory.xlsx')

sprd_sheet = my_file['Sheet1']

supplier_dictionary = {} # basically a hashmap
inventory_dictionary = {}

def company_count(row_count):
    for product_row in range(2, row_count + 1):
        supplier_name = sprd_sheet.cell(product_row,4).value

        if supplier_name in supplier_dictionary:
            supplier_dictionary[supplier_name] = supplier_dictionary[supplier_name] + 1
        else:
            supplier_dictionary[supplier_name] = 1 # key is name and sets initial value to 1

    print(supplier_dictionary)



def total_inventory_price_company(row_count):
    for product_row in range(2, row_count + 1):
        supplier_name = sprd_sheet.cell(product_row,4).value
        price = sprd_sheet.cell(product_row,3).value
        inventory = sprd_sheet.cell(product_row,2).value

        if supplier_name in inventory_dictionary:
            inventory_dictionary[supplier_name] = inventory_dictionary[supplier_name] + price * inventory
        else:
            inventory_dictionary[supplier_name] = inventory * price

    print(inventory_dictionary)


company_count(sprd_sheet.max_row)
total_inventory_price_company(sprd_sheet.max_row)